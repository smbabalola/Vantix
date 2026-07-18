"""Frozen-payload-only PDF and Excel renderers."""

from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass
from datetime import UTC, datetime
from html import escape
from importlib.metadata import version
from io import BytesIO
from typing import Any, Literal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from vantix_core.canonical import canonical_bytes


@dataclass(frozen=True, slots=True)
class RenderedArtefact:
    format: Literal["pdf", "xlsx"]
    content: bytes
    binary_checksum: str
    payload_checksum: str
    template_version: str
    renderer_version: str


class RendererUnavailable(RuntimeError):
    pass


def _binary_checksum(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def render_pdf(payload: dict[str, Any], payload_checksum: str) -> RenderedArtefact:
    """Render a compact foundation report without querying or recalculating live data."""

    try:
        from weasyprint import HTML  # type: ignore[import-untyped]
    except (ImportError, OSError) as exc:
        raise RendererUnavailable("WeasyPrint system libraries are unavailable.") from exc

    report = payload["report"]
    revision = payload["revision"]
    operations = payload.get("operations", {})
    rows = "".join(
        f"<tr><th>{escape(str(key).replace('_', ' ').title())}</th>"
        f"<td><pre>{escape(json.dumps(value, ensure_ascii=False, indent=2))}</pre></td></tr>"
        for key, value in sorted(operations.items())
    )
    html = f"""
    <!doctype html><html><head><meta charset="utf-8"><style>
      @page {{ size: A4; margin: 16mm; @bottom-right {{ content: "Page " counter(page); }} }}
      body {{ font-family: sans-serif; color: #12242b; font-size: 10pt; }}
      h1 {{ color: #0d635c; margin-bottom: 2mm; }}
      .meta {{ color: #53676c; margin-bottom: 8mm; }}
      table {{ width: 100%; border-collapse: collapse; }}
      th, td {{ border: 1px solid #d8e0df; padding: 3mm; vertical-align: top; }}
      th {{ width: 28%; text-align: left; background: #eef6f4; }}
      pre {{ white-space: pre-wrap; margin: 0; font-family: sans-serif; }}
      footer {{ margin-top: 8mm; color: #66777b; font-size: 8pt; }}
    </style></head><body>
      <h1>Vantix Daily Fluids Report</h1>
      <div class="meta">Report {escape(str(report["number"]))} · {escape(str(report["date"]))}<br>
      Revision {revision["number"]} · {escape(str(revision["state"]))}</div>
      <table>{rows}</table>
      <footer>Payload checksum: {escape(payload_checksum)}</footer>
    </body></html>
    """
    content = HTML(string=html).write_pdf()
    return RenderedArtefact(
        "pdf",
        content,
        _binary_checksum(content),
        payload_checksum,
        str(payload["template"]["version"]),
        f"weasyprint-{version('weasyprint')}",
    )


def render_excel(payload: dict[str, Any], payload_checksum: str) -> RenderedArtefact:
    """Render stable workbook structure from the same frozen canonical payload."""

    workbook = Workbook()
    active_sheet = workbook.active
    if active_sheet is None:
        raise RuntimeError("New workbook did not contain an active worksheet.")
    workbook.remove(active_sheet)
    workbook.properties.creator = "Vantix"
    workbook.properties.title = "Daily Fluids Report"
    fixed_timestamp = datetime(2000, 1, 1, tzinfo=UTC).replace(tzinfo=None)
    workbook.properties.created = fixed_timestamp
    workbook.properties.modified = fixed_timestamp

    summary = workbook.create_sheet("Summary")
    summary.append(["Vantix Daily Fluids Report", ""])
    summary.append(["Report number", payload["report"]["number"]])
    summary.append(["Report date", payload["report"]["date"]])
    summary.append(["Revision", payload["revision"]["number"]])
    summary.append(["State", payload["revision"]["state"]])
    summary.append(["Payload checksum", payload_checksum])
    summary["A1"].font = Font(bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="147D75")
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 72

    operations = workbook.create_sheet("Operations")
    operations.append(["Section", "Frozen payload value"])
    for key, value in sorted(payload.get("operations", {}).items()):
        operations.append([key, json.dumps(value, ensure_ascii=False, sort_keys=True)])
    for cell in operations[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DFF2EE")
    operations.column_dimensions["A"].width = 24
    operations.column_dimensions["B"].width = 100

    metadata = workbook.create_sheet("Audit Metadata")
    metadata.sheet_state = "visible"
    metadata.append(["payload_checksum", payload_checksum])
    metadata.append(["payload_schema_version", payload["schema_version"]])
    metadata.append(["template_key", payload["template"]["key"]])
    metadata.append(["template_version", payload["template"]["version"]])
    metadata.append(["canonical_payload", canonical_bytes(payload).decode("utf-8")])

    stream = BytesIO()
    workbook.save(stream)
    content = stream.getvalue()
    return RenderedArtefact(
        "xlsx",
        content,
        _binary_checksum(content),
        payload_checksum,
        str(payload["template"]["version"]),
        f"openpyxl-{version('openpyxl')}",
    )


def render_report(
    format: Literal["pdf", "xlsx"], payload: dict[str, Any], payload_checksum: str
) -> RenderedArtefact:
    if format == "pdf":
        return render_pdf(payload, payload_checksum)
    return render_excel(payload, payload_checksum)
