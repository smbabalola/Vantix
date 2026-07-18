from io import BytesIO

import pytest
from app.renderers import RendererUnavailable, render_excel, render_pdf
from openpyxl import load_workbook
from pypdf import PdfReader
from vantix_core.canonical import payload_checksum


def frozen_payload() -> dict:
    return {
        "schema_version": "1.0",
        "template": {"key": "daily-fluids-report", "version": "1.0"},
        "revision": {"number": 1, "state": "approved"},
        "report": {"number": "VTX-0001", "date": "2026-07-18"},
        "operations": {"general": {"operation_mode": "drilling"}},
    }


def test_vtx_det_005_006_pdf_and_excel_render_same_frozen_payload() -> None:
    payload = frozen_payload()
    checksum = payload_checksum(payload)
    try:
        pdf = render_pdf(payload, checksum)
    except RendererUnavailable:
        pytest.skip("WeasyPrint native libraries are provided by the backend container.")
    excel = render_excel(payload, checksum)
    assert pdf.content.startswith(b"%PDF")
    assert pdf.payload_checksum == excel.payload_checksum == checksum
    assert pdf.binary_checksum != excel.binary_checksum


def test_vtx_rpt_009_excel_embeds_authoritative_payload_checksum() -> None:
    payload = frozen_payload()
    checksum = payload_checksum(payload)
    artefact = render_excel(payload, checksum)
    workbook = load_workbook(BytesIO(artefact.content), read_only=True)
    metadata = workbook["Audit Metadata"]
    assert metadata["B1"].value == checksum
    assert workbook.sheetnames == ["Summary", "Operations", "Audit Metadata"]


def test_vtx_rpt_009_pdf_metadata_uses_unicode_separator_without_mojibake() -> None:
    payload = frozen_payload()
    try:
        artefact = render_pdf(payload, payload_checksum(payload))
    except RendererUnavailable:
        pytest.skip("WeasyPrint native libraries are provided by the backend container.")
    text = "\n".join(
        page.extract_text() or "" for page in PdfReader(BytesIO(artefact.content)).pages
    )
    assert "Â" not in text
    assert "VTX-0001" in text
    assert "2026-07-18" in text
